from rest_framework.viewsets import GenericViewSet, ModelViewSet
from rest_framework.mixins import CreateModelMixin, ListModelMixin, RetrieveModelMixin, DestroyModelMixin, UpdateModelMixin
from drf_spectacular.utils import extend_schema, extend_schema_view
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.tokens import UntypedToken
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.hashers import make_password

import logging
from openpyxl import load_workbook

import time

logger = logging.getLogger('main')

from general.serializers import (
    UserRegistrationSerializer, 
    UserListGeneralSerializer, 
    UserRetrieveSerializer,
    UserImportSerializer,
    # CustomTokenObtainPairSerializer,
    ClassGroupListGeneralSerializer,
    ClassRetrieveSerializer,
)

from general.models import (
    User
)

from general.services import (
    get_user_queryset,
    get_group_queryset
)

@extend_schema_view(post=extend_schema(summary='Получение токена', tags=['База: Аутентификация']))
class CustomTokenObtainPairView(TokenObtainPairView):
    # serializer_class = CustomTokenObtainPairSerializer
    def post(self, request, *args, **kwargs):
        # Вызов родительского метода для генерации токенов
        response = super().post(request, *args, **kwargs)

        # Получение токена доступа из ответа
        access_token = response.data.get("access", None)
        if access_token:
            try:
                # Расшифровка токена
                untyped_token = UntypedToken(access_token)

                # Получение ID пользователя из токена
                user_id = untyped_token['user_id']

                # Получение объекта пользователя
                user = User.objects.get(id=user_id)
                
                # Теперь у вас есть объект пользователя, можно добавить дополнительные данные в ответ, если нужно
                # Например:
                # response.data['username'] = user.username
                logger.info(f"User {user.email} (ID: {user.pk}) has been authenticated")

            except (InvalidToken, TokenError, User.DoesNotExist) as e:
                # Обработка ошибок, если токен невалиден или пользователя не существует
                logger.info(f"Authentication failed")
            
        return response

@extend_schema_view(post=extend_schema(summary='Обновление токена',tags=['База: Аутентификация']))
class CustomTokenRefreshView(TokenRefreshView):
    def post(self, request, *args, **kwargs):
        # Вызов родительского метода для обновления токенов
        response = super().post(request, *args, **kwargs)
        refresh_token = request.data.get("refresh")
        if refresh_token:
            try:
                decoded_token = RefreshToken(refresh_token)
                # Получите время истечения токена
                exp_timestamp = decoded_token['exp']
                current_timestamp = int(time.time())
                remaining_lifetime = exp_timestamp - current_timestamp
                logger.info(f"Refresh token lifetime is {remaining_lifetime} seconds")
            except TokenError:
                # Обработка ошибки расшифровки токена
                pass
        return response

@extend_schema_view(
    list=extend_schema(summary='Вывод списка пользователей', tags=['База: Пользователи']),
    create=extend_schema(summary='Создание нового пользователя', tags=['База: Пользователи']),
    retrieve=extend_schema(summary='Вывод информации о пользователе', tags=['База: Пользователи']),
    update=extend_schema(summary='Обновление информации о пользователе', tags=['База: Пользователи']),
    partial_update=extend_schema(summary='Частичное обновление информации о пользователе', tags=['База: Пользователи']),
    me=extend_schema(summary='Вывод информации об авторизованном пользователе', tags=['База: Пользователи']),
    user_import=extend_schema(summary='Импорт пользователей', tags=['База: Пользователи']),
    upload_photo=extend_schema(summary='Загрузка фотографии', tags=['База: Пользователи']),
    )
class UserViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return get_user_queryset()
    
    # переопределение сериализатора в зависимости от action
    def get_serializer_class(self):
        if self.action == "create":
            return UserRegistrationSerializer
        if self.action in ["retrieve", "me"]:
            return UserRetrieveSerializer
        return UserListGeneralSerializer
    
    @action(detail=False, methods=["get"], url_path="me")
    def me(self, request):
        instance = self.request.user
        serializer = self.get_serializer(instance)
        # Получаем заголовок Authorization
        auth_header = request.META.get('HTTP_AUTHORIZATION')
        if auth_header:
            # Разделяем 'Bearer' и токен
            parts = auth_header.split()
            if len(parts) == 2 and parts[0].lower() == "bearer":
                token = parts[1]
                try:
                    untyped_token = UntypedToken(token)
                    exp_timestamp = untyped_token['exp']
                    current_timestamp = int(time.time())
                    remaining_time = exp_timestamp - current_timestamp
                    logger.info(f"Token lifetime for {instance.email} (ID: {instance.pk}) is {remaining_time} seconds")
                except InvalidToken:
                    # Обработка невалидного токена
                    pass
            else:
                # Обработка ошибки: неверный формат заголовка
                pass
        else:
            # Обработка ошибки: заголовок отсутствует
            pass

        return Response(serializer.data)
    
    @action(detail=True, methods=["post"], url_path="upload")
    def upload_photo(self, request, pk=None):
        photo_file = request.FILES.get('photo')

        if photo_file is None:
            return Response({'detail': 'Invalid file'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.get(id=pk)
        except Exception as e:
            logger.error(f"An error occurred: {str(e)}")
            return Response({'detail': f'Error get user: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        user.photo = photo_file
        user.save()

        try:
            response_user_photo = User.objects.get(id=pk)
        except Exception as e:
            logger.error(f"An error occurred: {str(e)}")
        logger.info(f"User {user.email} (ID: {user.pk}) uploaded photo")
        return Response({'detail': 'Users photo success uploaded', 'user': self.get_serializer(user).data}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="import")
    def user_import(self, request):

        # Проверяем, есть ли файл формата .xlsx в POST-запросе
        file = request.FILES.get('file')
        if file is None or not file.name.endswith('.xlsx'):
            return Response({'detail': 'Invalid file format. Please upload an Excel file.'}, status=status.HTTP_400_BAD_REQUEST)

        # Загружаем файл в workbook с помощью openpyxl
        try:
            workbook = load_workbook(file)
            sheet = workbook.active
        except Exception as e:
            logger.error(f"An error occurred: {str(e)}")
            return Response({'detail': f'Error loading Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Создаём массив заголовков таблицы
        header = [cell.value for cell in sheet[1]]
        
        # Создаём массив словарей из таблицы
        user_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item_dict = dict(zip(header, row))
            item_dict['groups'] = str(item_dict['groups']).replace(".", ",").split(',')
            item_dict['gender'] = item_dict['gender'].strip()
            if 'departments' in item_dict:
                if item_dict['departments'] == None:
                    item_dict.pop('departments')
                else:
                    item_dict['departments'] = str(item_dict['departments']).replace(".", ",").split(',')
            if 'classes' in item_dict:
                if item_dict['classes'] == None:
                    item_dict.pop('classes')
                else:
                    item_dict['classes'] = str(item_dict['classes']).replace(".", ",").split(',')
            user_data.append(item_dict)

        # Сериализуем данные и выплоняем импорт в базу данных
        serializer = UserImportSerializer(data=user_data, many=True)
        if serializer.is_valid():
            # Собраем массив экземпляров модели пользователей
            users_to_create = []
            for user_data in serializer.validated_data:
                new_user = User(
                    email=user_data['email'],
                    password=make_password(user_data['password']),
                    first_name=user_data['first_name'],
                    middle_name=user_data.get('middle_name', None),
                    last_name=user_data['last_name'],
                    gender=user_data.get('gender', None),
                    position=user_data.get('position', None),
                    dnevnik_id=user_data.get('dnevnik_id', None),
                    dnevnik_user_id=user_data.get('dnevnik_user_id', None),
                )
                users_to_create.append(new_user)
            try:
                # Выполняем массовое создание пользователей из массива экземпляров
                users_instances = User.objects.bulk_create(users_to_create)
                # Выполняем добавление к созданным пользователям значения полей ManyToMany
                for user in users_instances:
                    # Поиск пользователя из сериализатора с таким же email, как текущего созданного пользователя
                    current_data = list(filter(lambda item: item['email'] == user.email, serializer.validated_data))[0]
                    # Добавление в поля ManyToMany значений id из полей сериализатора
                    user.departments.set(current_data.get('departments', []))
                    user.groups.set(current_data.get('groups', []))
                    user.classes.set(current_data.get('classes', []))
            except Exception as e:
                logger.error(f"An error occurred: {str(e)}")
                # Если произошла ошибка, удаляем созданных пользователей
                for user_instance in users_instances:
                    user_instance.delete()
                return Response({'detail': 'Users not imported'}, status=status.HTTP_400_BAD_REQUEST)
            logger.info(f"User {request.user.email} (ID: {request.user.id}) perform import users")
            return Response({'detail': 'Users imported successfully'}, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@extend_schema_view(
    list=extend_schema(summary='Список классов', tags=['База: Классы']),
    retrieve=extend_schema(summary='Просмотр класса со студентами', tags=['База: Классы']),
    )
class ClassGroupViewSet(ListModelMixin, RetrieveModelMixin, GenericViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return get_group_queryset()
    
    def get_serializer_class(self):
        if self.action in ["retrieve"]:
            return ClassRetrieveSerializer
        return ClassGroupListGeneralSerializer
    
    